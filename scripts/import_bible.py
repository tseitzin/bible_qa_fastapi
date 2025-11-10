#!/usr/bin/env python3
"""Load the KJV bible from an Excel workbook into the bible_verses table."""
import argparse
import logging
import sys
from itertools import chain
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Tuple

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# Ensure the backend package is importable when the script is run directly.
BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.config import get_settings  # noqa: E402

LOGGER = logging.getLogger(__name__)
EXPECTED_COLUMNS = {"book", "chapter", "verse", "text"}
DEFAULT_BATCH_SIZE = 1000
DEFAULT_EXCEL_PATH = Path(__file__).resolve().parents[2] / "kjv.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import bible verses from an Excel workbook")
    parser.add_argument(
        "--excel-path",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help="Path to the Excel file with KJV data (default: %(default)s)",
    )
    parser.add_argument(
        "--sheet-name",
        type=str,
        default=None,
        help="Worksheet name to load (defaults to the active sheet)",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help="Number of rows to insert per batch",
    )
    parser.add_argument(
        "--no-truncate",
        action="store_true",
        help="Skip truncating bible_verses before loading",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse the workbook without writing to the database",
    )
    return parser.parse_args()


def get_column_map(header_row: Iterable) -> Dict[str, int]:
    column_map: Dict[str, int] = {}
    for idx, value in enumerate(header_row):
        if value is None:
            continue
        key = str(value).strip().lower()
        if key in EXPECTED_COLUMNS and key not in column_map:
            column_map[key] = idx
    missing = EXPECTED_COLUMNS - set(column_map)
    if missing:
        raise ValueError(f"Missing expected columns: {', '.join(sorted(missing))}")
    return column_map


def iter_structured_rows(rows_iter: Iterator, column_map: Dict[str, int]) -> Iterator[Tuple[str, int, int, str]]:
    for row in rows_iter:
        if row is None:
            continue
        try:
            raw_book = row[column_map["book"]]
            raw_chapter = row[column_map["chapter"]]
            raw_verse = row[column_map["verse"]]
            raw_text = row[column_map["text"]]
        except IndexError as exc:
            raise ValueError("Row does not contain all required columns") from exc
        if raw_book is None or raw_text is None:
            continue
        book = str(raw_book).strip()
        if not book:
            continue
        try:
            chapter = int(str(raw_chapter).strip())
            verse = int(str(raw_verse).strip())
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid chapter/verse values for book '{book}'") from exc
        text = str(raw_text).strip()
        if not text:
            continue
        yield (book, chapter, verse, text)


def parse_reference_cell(reference: str) -> Tuple[str, int, int]:
    parts = reference.rsplit(" ", 1)
    if len(parts) != 2:
        raise ValueError("Reference must include book name and chapter/verse")
    book = parts[0].strip()
    chapter_verse = parts[1].strip()
    if ":" not in chapter_verse:
        raise ValueError("Chapter and verse separator ':' missing")
    chapter_str, verse_str = chapter_verse.split(":", 1)
    chapter_str = chapter_str.strip()
    verse_str = verse_str.strip().split("-", 1)[0]
    return book, int(chapter_str), int(verse_str)


def iter_reference_rows(rows_iter: Iterator) -> Iterator[Tuple[str, int, int, str]]:
    for row in rows_iter:
        if not row:
            continue
        reference = row[0]
        text = row[1] if len(row) > 1 else None
        if reference is None or text is None:
            continue
        reference_str = str(reference).strip()
        text_str = str(text).strip()
        if not reference_str or not text_str:
            continue
        try:
            book, chapter, verse = parse_reference_cell(reference_str)
        except ValueError:
            continue
        yield (book, chapter, verse, text_str)


def load_workbook(path: Path, sheet_name: str | None):
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    workbook = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
        return workbook[sheet_name]
    return workbook.active


def load_data(args: argparse.Namespace) -> List[Tuple[str, int, int, str]]:
    sheet = load_workbook(args.excel_path, args.sheet_name)
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        first_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Worksheet is empty") from exc
    try:
        column_map = get_column_map(first_row)
    except ValueError:
        LOGGER.info("Structured header not detected; falling back to reference parsing")
        rows_iter = chain([first_row], rows_iter)
        return list(iter_reference_rows(rows_iter))
    return list(iter_structured_rows(rows_iter, column_map))


def insert_rows(rows: List[Tuple[str, int, int, str]], batch_size: int, dry_run: bool, truncate: bool) -> None:
    settings = get_settings()
    if dry_run:
        LOGGER.info("Dry run enabled; skipping database writes")
        LOGGER.info("Would import %d rows", len(rows))
        return
    with psycopg2.connect(**settings.db_config) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT to_regclass('public.bible_verses');")
            exists = cur.fetchone()[0]
            if exists is None:
                raise RuntimeError("Table bible_verses does not exist. Run migrations before importing.")
            cur.execute("BEGIN;")
            if truncate:
                LOGGER.info("Clearing existing bible_verses data")
                cur.execute("TRUNCATE TABLE bible_verses RESTART IDENTITY;")
            else:
                LOGGER.info("Preserving existing bible_verses data; duplicates will be updated")
            total = 0
            sql = (
                "INSERT INTO bible_verses (book, chapter, verse, text) VALUES %s "
                "ON CONFLICT (book, chapter, verse) DO UPDATE SET text = EXCLUDED.text"
            )
            for start in range(0, len(rows), batch_size):
                chunk = rows[start:start + batch_size]
                execute_values(cur, sql, chunk, page_size=len(chunk))
                total += len(chunk)
                LOGGER.info("Inserted %d/%d rows", total, len(rows))
            conn.commit()
    LOGGER.info("Import complete: %d rows inserted", len(rows))


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    args = parse_args()
    LOGGER.info("Loading workbook from %s", args.excel_path)
    rows = load_data(args)
    LOGGER.info("Parsed %d rows from the workbook", len(rows))
    if not rows:
        LOGGER.warning("Workbook did not contain any verse rows; exiting")
        return
    insert_rows(rows, args.batch_size, args.dry_run, not args.no_truncate)


if __name__ == "__main__":
    main()
